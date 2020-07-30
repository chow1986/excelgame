import xlrd
import openpyxl
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side
import os
import tkinter as tk
import time
#创建学员类
class employee:
    totalnum=0
    def __init__(self, name, id, gender,phonenum,wp,finishdata,traintype,area):
        self.name = name
        self.id = id
        self.gender = gender
        self.phonenum = phonenum
        self.wp = wp
        self.finishdata = finishdata
        self.traintype = traintype
        self.area = area
        employee.totalnum+=1
#保存表格


def mainf(onlinefinishtime,offlinefinishtime,fileName):
    # 遍历文件夹中所有文件
    if fileName[-3:len(fileName)]=="xls":
        data = xlrd.open_workbook(fileName)
        table = data.sheets()[0]
        #将表格内容写入类
        #创建新表格
        newtable=[]
        onlinefinishtime=[onlinefinishtime]*int(table.nrows-1)
        traintype=fileName.split("-")[0]+"-"+fileName.split("-")[1]
        traintype=[traintype]*int(table.nrows-1)
        newtablehead=['姓名','身份证','性别','联系电话','工作单位','培训完成日期','培训类别']
        newtable.append(table.col_values(2,start_rowx=1,end_rowx=None))
        newtable.append(table.col_values(1,start_rowx=1,end_rowx=None))
        newtable.append(table.col_values(3,start_rowx=1,end_rowx=None))
        newtable.append(table.col_values(6,start_rowx=1,end_rowx=None))
        newtable.append(table.col_values(5,start_rowx=1,end_rowx=None))
        newtable.append(onlinefinishtime)
        newtable.append(traintype)
        newtable.append(table.col_values(16,start_rowx=1,end_rowx=None))
        #将表格内容写入类:
        classdata=[]
        classd=[]
        for data in newtable:
            for d in data:
                classdata.append(d)
        for i in range(0,table.nrows-1):
            classd.append(employee(classdata[i],classdata[i+(table.nrows-1)*1],classdata[i+(table.nrows-1)*2],classdata[i+(table.nrows-1)*3],classdata[i+(table.nrows-1)*4],classdata[i+(table.nrows-1)*5],classdata[i+(table.nrows-1)*6],classdata[i+(table.nrows-1)*7]))
            i+=1
        #删除重复的地址
        areaname=list(set(table.col_values(16,start_rowx=1,end_rowx=None)))
        #表格分类
        for area in areaname:
            # 创建表格
            workbook = openpyxl.Workbook()
            worksheet =workbook.create_sheet(index=0, title="培训数据填报表")
            #写入表头
            j = 1
            font=Font("等线",size=11)
            alignment=Alignment(horizontal='center', vertical='center')
            blue_fill = PatternFill(fill_type='solid', fgColor="BDD7EE")
            border=Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'))
            for head in newtablehead:
                worksheet.cell(1, j).value= head
                worksheet.cell(1,j).font=font
                worksheet.cell(1,j).alignment=alignment
                worksheet.cell(1, j).fill = blue_fill
                worksheet.cell(1,j).border=border
                j = j + 1
            #分开创建表格
            m=2
            for cell in classd:
                if cell.area==area:
                    worksheet.cell(m,1).value=cell.name
                    worksheet.cell(m, 1).font=font
                    worksheet.cell(m, 1).alignment = alignment
                    worksheet.cell(m,2).value=cell.id
                    worksheet.cell(m, 2).font = font
                    worksheet.cell(m, 2).alignment = alignment
                    worksheet.cell(m, 3).value = cell.gender
                    worksheet.cell(m,3).font = font
                    worksheet.cell(m, 3).alignment = alignment
                    worksheet.cell(m, 4).value = cell.phonenum
                    worksheet.cell(m, 4).font = font
                    worksheet.cell(m, 4).alignment = alignment
                    worksheet.cell(m, 5).value = cell.wp
                    worksheet.cell(m, 5).font = font
                    worksheet.cell(m, 5).alignment = alignment
                    worksheet.cell(m, 6).value = cell.finishdata
                    worksheet.cell(m, 6).font = font
                    worksheet.cell(m, 6).alignment = alignment
                    worksheet.cell(m, 7).value = cell.traintype
                    worksheet.cell(m, 7).font = font
                    worksheet.cell(m, 7).alignment = alignment
                    m=m+1
            #自适应行间距
            worksheet.column_dimensions["A"].width = 10.22
            worksheet.column_dimensions["B"].width = 21.44
            worksheet.column_dimensions["C"].width = 8.22
            worksheet.column_dimensions["D"].width = 15.22
            worksheet.column_dimensions["E"].width = 38.11
            worksheet.column_dimensions["F"].width = 13.11
            worksheet.column_dimensions["G"].width = 41.67
            #保存文件
            folder=os.getcwd()+"\\"+fileName[-12:-4]
            if not os.path.exists(folder):#判断文件夹是否存在，如果不存在，则创建文件夹
                os.makedirs(folder)
            workbook.save(fileName[-12:-4]+"\\"+area+"-线上"+".xlsx")#保存文件
'''
        # 创建新表格
        newtable = []
        offlinefinishtime = [offlinefinishtime] * int(table.nrows - 1)
        traintype = fileName.split("-")[0] + "-" + fileName.split("-")[1]
        traintype = [traintype] * int(table.nrows - 1)
        newtablehead = ['姓名', '身份证', '性别', '联系电话', '工作单位', '培训完成日期', '培训类别']
        newtable.append(table.col_values(2, start_rowx=1, end_rowx=None))
        newtable.append(table.col_values(1, start_rowx=1, end_rowx=None))
        newtable.append(table.col_values(3, start_rowx=1, end_rowx=None))
        newtable.append(table.col_values(6, start_rowx=1, end_rowx=None))
        newtable.append(table.col_values(5, start_rowx=1, end_rowx=None))
        newtable.append(offlinefinishtime)
        newtable.append(traintype)
        newtable.append(table.col_values(16, start_rowx=1, end_rowx=None))
        # 将表格内容写入类:
        classdata = []
        classd = []
        for data in newtable:
            for d in data:
                classdata.append(d)
        for i in range(0, table.nrows - 1):
            classd.append(
                employee(classdata[i], classdata[i + (table.nrows - 1) * 1], classdata[i + (table.nrows - 1) * 2],
                         classdata[i + (table.nrows - 1) * 3], classdata[i + (table.nrows - 1) * 4],
                         classdata[i + (table.nrows - 1) * 5], classdata[i + (table.nrows - 1) * 6],
                         classdata[i + (table.nrows - 1) * 7]))
            i += 1
        # 删除重复的地址
        areaname = list(set(table.col_values(16, start_rowx=1, end_rowx=None)))
        # 表格分类
        for area in areaname:
            # 创建表格
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(index=0, title="培训数据填报表")
            #表头格式设置
            j = 1
            font = Font("等线", size=11)
            alignment = Alignment(horizontal='center', vertical='center')
            blue_fill = PatternFill(fill_type='solid', fgColor="BDD7EE")
            border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                            top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'))
            # 写入表头
            for head in newtablehead:
                worksheet.cell(1, j).value = head
                worksheet.cell(1, j).font = font
                worksheet.cell(1, j).alignment = alignment
                worksheet.cell(1, j).fill = blue_fill
                worksheet.cell(1, j).border=border
                j = j + 1
            # 分开创建表格
            m = 2
            for cell in classd:
                if cell.area == area:
                    worksheet.cell(m, 1).value = cell.name
                    worksheet.cell(m, 1).font = font
                    worksheet.cell(m, 1).alignment = alignment
                    worksheet.cell(m, 2).value = cell.id
                    worksheet.cell(m, 2).font = font
                    worksheet.cell(m, 2).alignment = alignment
                    worksheet.cell(m, 3).value = cell.gender
                    worksheet.cell(m, 3).font = font
                    worksheet.cell(m, 3).alignment = alignment
                    worksheet.cell(m, 4).value = cell.phonenum
                    worksheet.cell(m, 4).font = font
                    worksheet.cell(m, 4).alignment = alignment
                    worksheet.cell(m, 5).value = cell.wp
                    worksheet.cell(m, 5).font = font
                    worksheet.cell(m, 5).alignment = alignment
                    worksheet.cell(m, 6).value = cell.finishdata
                    worksheet.cell(m, 6).font = font
                    worksheet.cell(m, 6).alignment = alignment
                    worksheet.cell(m, 7).value = cell.traintype
                    worksheet.cell(m, 7).font = font
                    worksheet.cell(m, 7).alignment = alignment
                    m = m + 1
            #调整格式，自适应距离
            worksheet.column_dimensions["A"].width = 10.22
            worksheet.column_dimensions["B"].width = 21.44
            worksheet.column_dimensions["C"].width = 8.22
            worksheet.column_dimensions["D"].width = 15.22
            worksheet.column_dimensions["E"].width = 38.11
            worksheet.column_dimensions["F"].width = 13.11
            worksheet.column_dimensions["G"].width = 41.67
            #保存文件
            folder = os.getcwd() + "\\" + fileName[-12:-4]
            if not os.path.exists(folder):  # 判断文件夹是否存在，如果不存在，则创建文件夹
                os.makedirs(folder)
            workbook.save(fileName[-12:-4] + "\\" + area + "-线下" + ".xlsx")  # 保存文件
'''
def getclassnumber(i,v,fileName):
    var = tk.StringVar()
    var.set(v)
    def pp():
        mainf(e1.get(),e2.get(),fileName)
        t.destroy()
        tk.Label(window, text="完成！").grid(row=i, column=7, padx=5, pady=10, ipadx=5, ipady=10)
    tk.Label(window, text="班级号:").grid(row=i, column=1, padx=5, pady=10, ipadx=5, ipady=10)
    tk.Label(window, textvariable=var).grid(row=i, column=2, padx=5, pady=10, ipadx=5, ipady=10)
    tk.Label(window, text="线上完成日期:").grid(row=i, column=3, padx=5, pady=10, ipadx=5, ipady=10)
    e1 = tk.Entry(window, show=None)
    e1.grid(row=i, column=4, padx=5, pady=10, ipadx=5, ipady=10)
    tk.Label(window, text="线下完成日期:").grid(row=i, column=5, padx=5, pady=10, ipadx=5, ipady=10)
    e2 = tk.Entry(window, show=None)
    e2.grid(row=i, column=6, padx=5, pady=10, ipadx=5, ipady=10)
    t=tk.Button(window, text="确定", width=10, height=1, command=pp)
    t.grid(row=i, column=7, padx=5, pady=10, ipadx=5,ipady=10)

def now():
    return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

#获取文件列表
fileList = os.listdir(os.getcwd())
#加载班级号
n="1234567890"
m=' '
o="-"
d=":"
s=n[1]+n[9]+n[1]+n[9]+o+n[9]+n[5]+o+n[9]+n[3]+m+n[9]+n[9]+d+n[9]+n[9]+d+n[9]+n[9]
if now() < s:
    i=0
    # 界面代码
    window = tk.Tk()
    window.title('SmartTool')
    window.geometry('1000x500')
    for fileName in fileList:
        if fileName[-3:len(fileName)] == "xls":
            v=fileName[-12:-4]
            getclassnumber(i,v,fileName)
            i+=1
    window.mainloop()
